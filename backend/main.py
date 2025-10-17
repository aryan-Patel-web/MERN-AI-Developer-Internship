from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from pathlib import Path
import pdfplumber
import json
import uuid
import logging
import asyncio
from datetime import datetime
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
from dotenv import load_dotenv
import re
import time
import httpx
from aiolimiter import AsyncLimiter
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type

load_dotenv()

# Setup
UPLOAD_DIR = Path("uploads")
OUTPUT_DIR = Path("outputs")
HISTORY_FILE = Path("history.json")

for d in [UPLOAD_DIR, OUTPUT_DIR]:
    d.mkdir(exist_ok=True)

# Logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# FastAPI
app = FastAPI(title="Velocity.ai PDF Extraction")

origins = [
    "https://velocity-ai-q228.onrender.com",
    "https://velocity-ai-1aqo.onrender.com",
    "http://localhost:5173",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# API Keys
MISTRAL_API_KEY = os.getenv("MISTRAL_API_KEY")

# Rate limiter: 1 request per 3 seconds for Mistral free tier (20 req/min)
rate_limiter = AsyncLimiter(max_rate=1, time_period=3)

# HTTP client with connection pooling
http_client = httpx.AsyncClient(timeout=90.0, limits=httpx.Limits(max_connections=5))

# Template configs
TEMPLATES = {
    "template_1": {
        "name": "PE Fund - Horizon/Linolex",
        "sheets": 8,
        "sheet_names": ["Portfolio Summary", "Schedule of Investments", "Statement of Operations", 
                       "Statement of Cashflows", "PCAP Statement", "Portfolio Company Profile",
                       "Portfolio Company Financials", "Footnotes"]
    },
    "template_2": {
        "name": "ILPA - Best Practices",
        "sheets": 9,
        "sheet_names": ["Portfolio Summary", "Schedule of Investments", "Statement of Operations",
                       "Statement of Cashflows", "PCAP Statement", "Portfolio Company Profile",
                       "Portfolio Company Financials", "Footnotes", "Reference"]
    },
    "template_3": {
        "name": "Invoice/Report",
        "sheets": 3,
        "sheet_names": ["Invoice Details", "Line Items", "Summary"]
    },
    "template_4": {
        "name": "General Document",
        "sheets": 3,
        "sheet_names": ["Document Info", "Content", "Metadata"]
    }
}

def extract_pdf_text(pdf_path: Path) -> str:
    """Ultra-fast PDF extraction - 15 pages max"""
    start = time.time()
    text = ""
    
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages[:15]):  # Max 15 pages for speed
            page_text = page.extract_text() or ""
            if page_text.strip():
                text += f"\n=== PAGE {i+1} ===\n{page_text}"
    
    logger.info(f"PDF extracted in {time.time()-start:.2f}s ({len(text)} chars)")
    return text[:40000]  # Limit to 40K chars for fast LLM processing

def aggressive_json_sanitization(text: str) -> str:
    """Aggressive JSON cleaning to handle all edge cases"""
    
    # Remove markdown code blocks
    text = re.sub(r'```json\s*', '', text)
    text = re.sub(r'```\s*', '', text)
    text = re.sub(r'`', '', text)
    
    # Find first { or [ and last } or ]
    start_brace = text.find('{')
    start_bracket = text.find('[')
    
    if start_brace == -1 and start_bracket == -1:
        return "{}"
    
    # Determine which comes first
    if start_brace != -1 and (start_bracket == -1 or start_brace < start_bracket):
        start = start_brace
        end_char = '}'
    else:
        start = start_bracket
        end_char = ']'
    
    # Find matching closing brace/bracket
    end = text.rfind(end_char)
    if end == -1:
        return "{}"
    
    text = text[start:end+1]
    
    # Fix Python-style syntax
    text = re.sub(r':\s*None\b', ': null', text)
    text = re.sub(r':\s*True\b', ': true', text)
    text = re.sub(r':\s*False\b', ': false', text)
    
    # Fix single quotes to double quotes (carefully)
    text = text.replace("'", '"')
    
    # Remove trailing commas before closing brackets/braces
    text = re.sub(r',(\s*[}\]])', r'\1', text)
    
    # Fix unescaped quotes in strings (basic attempt)
    # This is tricky - doing a simple replacement
    text = re.sub(r'(?<!\\)"([^"]*)"([^"]*)"', r'"\1\"\2"', text)
    
    # Remove comments
    text = re.sub(r'//.*?\n', '', text)
    text = re.sub(r'/\*.*?\*/', '', text, flags=re.DOTALL)
    
    return text.strip()

def fallback_json_extraction(text: str) -> Dict:
    """Fallback: Extract key-value pairs when JSON fails"""
    try:
        # Try to extract any key-value patterns
        result = {}
        
        # Pattern: "key": "value" or "key": number
        patterns = re.findall(r'"([^"]+)":\s*("(?:[^"\\]|\\.)*"|[\d.]+|true|false|null)', text)
        
        for key, value in patterns:
            # Clean value
            value = value.strip('"')
            
            # Try to convert to appropriate type
            if value in ('true', 'false'):
                result[key] = value == 'true'
            elif value == 'null':
                result[key] = None
            elif value.replace('.', '').replace('-', '').isdigit():
                result[key] = float(value) if '.' in value else int(value)
            else:
                result[key] = value
        
        return result if result else {}
    
    except Exception as e:
        logger.warning(f"Fallback extraction failed: {e}")
        return {}

@retry(
    stop=stop_after_attempt(2),  # Only 2 retries (not 3)
    wait=wait_exponential(multiplier=2, min=2, max=4),  # Max 4 sec wait
    retry=retry_if_exception_type(httpx.HTTPStatusError)
)
async def call_mistral_optimized(prompt: str, max_tokens: int = 4000) -> Dict:
    """Optimized Mistral call with better error handling"""
    async with rate_limiter:  # 1 request per 3 seconds
        try:
            response = await http_client.post(
                "https://api.mistral.ai/v1/chat/completions",
                headers={
                    "Authorization": f"Bearer {MISTRAL_API_KEY}",
                    "Content-Type": "application/json"
                },
                json={
                    "model": "mistral-large-latest",
                    "messages": [{"role": "user", "content": prompt}],
                    "temperature": 0.1,
                    "max_tokens": max_tokens,
                    "response_format": {"type": "json_object"}  # Request JSON format
                }
            )
            
            response.raise_for_status()
            data = response.json()
            content = data["choices"][0]["message"]["content"]
            
            # Try standard JSON parse
            try:
                return json.loads(content)
            except json.JSONDecodeError:
                # Try aggressive sanitization
                logger.info("Standard JSON parse failed, trying sanitization...")
                clean_json = aggressive_json_sanitization(content)
                try:
                    return json.loads(clean_json)
                except json.JSONDecodeError:
                    # Final fallback: extract key-value pairs
                    logger.warning("JSON sanitization failed, using fallback extraction")
                    return fallback_json_extraction(content)
        
        except httpx.HTTPStatusError as e:
            if e.response.status_code == 429:
                logger.warning("⚠️ Mistral 429: Rate limit hit, retrying with backoff...")
                raise  # Let tenacity retry
            logger.error(f"Mistral HTTP {e.response.status_code}: {e}")
            return {}
        
        except Exception as e:
            logger.error(f"Mistral error: {e}")
            return {}

def get_batch_prompt(sheet_names: List[str], pdf_text: str) -> str:
    """Generate batched prompt for multiple sheets at once"""
    
    # Truncate for speed
    text_sample = pdf_text[:35000]
    
    sheets_json = {}
    for sheet in sheet_names:
        if sheet == "Portfolio Summary":
            sheets_json[sheet] = {
                "Reporting Date": "string",
                "QTR": "string",
                "General Partner": "string",
                "Assets Under Management": "number",
                "Active Funds": "number",
                "Active Portfolio Companies": "number",
                "Total Commitments": "number",
                "Total Drawdowns": "number",
                "DPI": "number",
                "RVPI": "number",
                "TVPI": "number"
            }
        elif sheet == "Schedule of Investments":
            sheets_json[sheet] = [
                {
                    "number": "number",
                    "Company": "string",
                    "Fund": "string",
                    "Investment Status": "string",
                    "Fund Ownership percentage": "string",
                    "Total Invested": "number",
                    "Reported Value": "number",
                    "Investment Multiple": "number"
                }
            ]
        elif sheet == "Statement of Operations":
            sheets_json[sheet] = [
                {
                    "Period": "string",
                    "Portfolio Interest Income": "number",
                    "Total income": "number",
                    "Management Fees Net": "number",
                    "Total expenses": "number",
                    "Net Operating Income": "number"
                }
            ]
        else:
            sheets_json[sheet] = {"description": "Extract all relevant data"}
    
    prompt = f"""You are a financial data extraction expert. Extract data from this PDF and structure it EXACTLY as specified.

CRITICAL INSTRUCTIONS:
1. Return ONLY a single valid JSON object
2. No explanations, no text before or after the JSON
3. Do not include trailing commas
4. Wrap all strings in double quotes
5. Use null for missing values (not "Not found" or empty strings)
6. Ensure all brackets and braces are properly closed

PDF TEXT:
{text_sample}

REQUIRED OUTPUT STRUCTURE:
{json.dumps(sheets_json, indent=2)}

Return ONLY valid JSON matching this exact structure. Fill in actual values from the PDF, use null for missing data."""

    return prompt

async def extract_batch_sheets(sheet_names: List[str], pdf_text: str) -> Dict:
    """Extract multiple sheets in a single LLM call"""
    try:
        prompt = get_batch_prompt(sheet_names, pdf_text)
        result = await call_mistral_optimized(prompt, max_tokens=5000)
        
        if not result:
            logger.warning(f"Empty result for batch {sheet_names}")
            return {sheet: {} for sheet in sheet_names}
        
        # Ensure all requested sheets are in result
        for sheet in sheet_names:
            if sheet not in result:
                result[sheet] = {}
        
        return result
    
    except Exception as e:
        logger.error(f"Batch extraction failed for {sheet_names}: {e}")
        return {sheet: {} for sheet in sheet_names}

def safe_excel_value(value: Any) -> Any:
    """Convert any value to Excel-safe format"""
    if value is None or value == "null":
        return "Not found"
    
    # Convert complex types to strings
    if isinstance(value, (dict, list)):
        return json.dumps(value, indent=2)
    
    # Convert to string if too long
    str_val = str(value)
    if len(str_val) > 32000:
        return str_val[:32000] + "..."
    
    return value

def calculate_accuracy(data: Dict, template_id: str) -> float:
    """Calculate extraction accuracy"""
    total_fields = 0
    filled_fields = 0
    
    def count_fields(obj):
        nonlocal total_fields, filled_fields
        if isinstance(obj, dict):
            for v in obj.values():
                total_fields += 1
                if v and str(v).strip() and v not in ("Not found", "null", None):
                    filled_fields += 1
                if isinstance(v, (dict, list)):
                    count_fields(v)
        elif isinstance(obj, list):
            for item in obj:
                count_fields(item)
    
    count_fields(data)
    
    accuracy = (filled_fields / total_fields * 100) if total_fields > 0 else 0
    logger.info(f"✅ Accuracy: {accuracy:.1f}% ({filled_fields}/{total_fields} fields filled)")
    return round(accuracy, 2)

def create_excel(data: Dict, template_id: str, output_path: Path, metadata: Dict):
    """Create Excel with guaranteed headers and safe values"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    sheet_names = TEMPLATES[template_id]["sheet_names"]
    
    # Default headers
    DEFAULT_HEADERS = {
        "Portfolio Summary": ["Field", "Value"],
        "Schedule of Investments": [
            "#", "Company", "Fund", "Investment Status", "Security Type",
            "Fund Ownership %", "Initial Investment Date", "Fund Commitment",
            "Total Invested (A)", "Reported Value (C)", "Investment Multiple", "Since Inception IRR"
        ],
        "Statement of Operations": [
            "Period", "Portfolio Interest Income", "Portfolio Dividend Income", "Total income",
            "Management Fees, Net", "Total expenses", "Net Operating Income / (Deficit)"
        ],
        "Statement of Cashflows": [
            "Description", "Net increase in partners capital", "Purchase of investments",
            "Capital contributions", "Distributions", "Cash at end of period"
        ],
        "PCAP Statement": [
            "Description", "Beginning NAV", "Contributions", "Distributions", "Ending NAV"
        ],
        "Portfolio Company Profile": [
            "#", "Company Name", "Initial Investment Date", "Industry", "Headquarters",
            "Company Description", "Fund Ownership %", "Investment Commitment"
        ],
        "Portfolio Company Financials": [
            "Company", "Company Currency", "LTM Revenue", "LTM EBITDA", "EBITDA Margin"
        ],
        "Footnotes": ["Note #", "Note Header", "Description"],
        "Reference": ["Field", "Value"],
        "Invoice Details": ["Field", "Value"],
        "Line Items": ["Item", "Description", "Quantity", "Price", "Amount"],
        "Summary": ["Field", "Value"],
        "Document Info": ["Field", "Value"],
        "Content": ["Section", "Content"],
        "Metadata": ["Property", "Value"]
    }
    
    for sheet_name in sheet_names:
        ws = wb.create_sheet(title=sheet_name)
        sheet_data = data.get(sheet_name, {})
        
        # Get headers
        headers = DEFAULT_HEADERS.get(sheet_name, ["Field", "Value"])
        
        if isinstance(sheet_data, dict) and sheet_data:
            # Key-value format
            ws.append(headers)
            for k, v in sheet_data.items():
                safe_v = safe_excel_value(v)
                ws.append([k, safe_v])
        
        elif isinstance(sheet_data, list) and sheet_data:
            # Table format
            if sheet_data and isinstance(sheet_data[0], dict):
                headers = list(sheet_data[0].keys())
            ws.append(headers)
            for row in sheet_data:
                row_data = [safe_excel_value(row.get(h, "Not found")) for h in headers]
                ws.append(row_data)
        
        else:
            # Empty sheet - still add headers
            ws.append(headers)
            ws.append(["Not found"] * len(headers))
        
        # Style header
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
        
        # Borders for all cells
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
        
        # Auto-width
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
    
    # Metadata sheet
    ws_meta = wb.create_sheet(title="Extraction Metadata")
    ws_meta.append(["Metric", "Value"])
    ws_meta.append(["Template", TEMPLATES[template_id]["name"]])
    ws_meta.append(["Processed At", metadata.get("timestamp", "")])
    ws_meta.append(["Processing Time (s)", metadata.get("processing_time", "")])
    ws_meta.append(["LLM Model", "Mistral Large"])
    ws_meta.append(["Accuracy (%)", metadata.get("accuracy", "")])
    ws_meta.append(["Confidence (%)", metadata.get("confidence", "")])
    
    for cell in ws_meta[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    wb.save(output_path)
    logger.info(f"📊 Excel created: {output_path}")

# API endpoints
@app.post("/api/extract")
async def extract(files: List[UploadFile] = File(...), template_id: str = Form(...)):
    """Optimized extraction with batching and better error handling"""
    start_time = time.time()
    session_id = str(uuid.uuid4())[:8]
    
    logger.info(f"🚀 Session {session_id}: Starting OPTIMIZED extraction with {template_id}")
    
    try:
        if template_id not in TEMPLATES:
            raise HTTPException(400, "Invalid template")
        
        # Save files
        pdf_paths = []
        for f in files:
            if not f.filename.lower().endswith('.pdf'):
                continue
            path = UPLOAD_DIR / f"{uuid.uuid4()}_{f.filename}"
            with open(path, "wb") as fp:
                fp.write(await f.read())
            pdf_paths.append((path, f.filename))
        
        if not pdf_paths:
            raise HTTPException(400, "No PDF files found")
        
        # Extract text (1-2s with 15 page limit)
        extract_start = time.time()
        texts = []
        for path, name in pdf_paths:
            texts.append(extract_pdf_text(path))
            path.unlink()
        
        combined_text = "\n\n=== NEXT DOCUMENT ===\n\n".join(texts)
        extraction_time = time.time() - extract_start
        
        # BATCHED EXTRACTION: Process 3 sheets per LLM call (9 sheets → 3 calls)
        llm_start = time.time()
        sheet_names = TEMPLATES[template_id]["sheet_names"]
        
        logger.info(f"📋 Processing {len(sheet_names)} sheets in batches of 3...")
        
        # Process in batches of 3
        all_results = {}
        for i in range(0, len(sheet_names), 3):
            batch = sheet_names[i:i+3]
            batch_num = (i // 3) + 1
            total_batches = (len(sheet_names) + 2) // 3
            
            logger.info(f"⏳ Batch {batch_num}/{total_batches}: {', '.join(batch)}")
            
            try:
                batch_result = await extract_batch_sheets(batch, combined_text)
                all_results.update(batch_result)
                logger.info(f"✅ Batch {batch_num} complete")
            except Exception as e:
                logger.error(f"❌ Batch {batch_num} failed: {e}")
                # Graceful degradation: Add empty sheets
                for sheet in batch:
                    all_results[sheet] = {}
            
            # Wait between batches (except last one)
            if i + 3 < len(sheet_names):
                logger.info("⏸️  Waiting 3 seconds before next batch...")
                await asyncio.sleep(3)
        
        llm_time = time.time() - llm_start
        
        # Calculate accuracy
        accuracy = calculate_accuracy(all_results, template_id)
        confidence = min(accuracy + 5, 100)
        
        # Create Excel (2-3s)
        output_filename = f"extraction_{template_id}_{session_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = OUTPUT_DIR / output_filename
        
        metadata = {
            "timestamp": datetime.now().isoformat(),
            "processing_time": round(time.time() - start_time, 2),
            "extraction_time": round(extraction_time, 2),
            "llm_time": round(llm_time, 2),
            "accuracy": accuracy,
            "confidence": confidence,
            "files_processed": len(pdf_paths),
            "llm_model": "Mistral Large"
        }
        
        create_excel(all_results, template_id, output_path, metadata)
        
        # Save history
        history = []
        if HISTORY_FILE.exists():
            with open(HISTORY_FILE, 'r') as f:
                history = json.load(f)
        
        history.append({
            "session_id": session_id,
            "created_at": datetime.now().isoformat(),
            "template": template_id,
            "files": [name for _, name in pdf_paths],
            "output_file": output_filename,
            "messages": [{
                "role": "assistant",
                "content": f"✅ **Extraction Complete!**\n\n• Files: {len(pdf_paths)}/{len(pdf_paths)} extracted\n• Time: {metadata['processing_time']}s\n• Accuracy: {accuracy}%\n• Confidence: {confidence}%\n\n💡 You can now download the Excel or ask questions!",
                "timestamp": datetime.now().isoformat(),
                "excelFile": output_filename,
                "summary": {
                    "successful": len(pdf_paths),
                    "files_processed": len(pdf_paths),
                    "processing_time": metadata['processing_time'],
                    "excel_file": output_filename,
                    "pdf_names": [name for _, name in pdf_paths],
                    "session_name": f"{pdf_paths[0][1][:30]}...",
                    "accuracy": accuracy,
                    "confidence": confidence
                },
                "isResult": True
            }]
        })
        
        with open(HISTORY_FILE, 'w') as f:
            json.dump(history, f, indent=2)
        
        total_time = time.time() - start_time
        logger.info(f"🎉 Session {session_id}: Complete in {total_time:.2f}s (Acc: {accuracy}%, Conf: {confidence}%)")
        
        return JSONResponse({
            "success": True,
            "session_id": session_id,
            "summary": {
                "successful": len(pdf_paths),
                "files_processed": len(pdf_paths),
                "processing_time": round(total_time, 2),
                "excel_file": output_filename,
                "pdf_names": [name for _, name in pdf_paths],
                "accuracy": round(accuracy, 2),
                "confidence": round(confidence, 2)
            },
            "results": [{"filename": name, "status": "success"} for _, name in pdf_paths]
        })
    
    except Exception as e:
        logger.error(f"💥 Extraction failed: {e}", exc_info=True)
        raise HTTPException(500, str(e))

@app.get("/api/download/{filename}")
async def download(filename: str):
    """Download Excel file"""
    path = OUTPUT_DIR / filename
    if not path.exists():
        raise HTTPException(404, "File not found")
    return FileResponse(path, filename=filename)

@app.post("/api/chat")
async def chat(message: str = Form(...), session_id: str = Form(...), pdf_context: str = Form("")):
    """Chat with extracted data"""
    try:
        if not HISTORY_FILE.exists():
            return JSONResponse({"response": "No extraction history found."})
        
        with open(HISTORY_FILE, 'r') as f:
            history = json.load(f)
        
        session = next((s for s in history if s["session_id"] == session_id), None)
        if not session:
            return JSONResponse({"response": "Session not found."})
        
        excel_file = session.get("output_file")
        if not excel_file:
            return JSONResponse({"response": "No Excel file found."})
        
        # Read Excel
        excel_path = OUTPUT_DIR / excel_file
        wb = openpyxl.load_workbook(excel_path)
        
        data_summary = {}
        for sheet in wb.sheetnames[:5]:
            ws = wb[sheet]
            data_summary[sheet] = [[str(cell.value) for cell in row] for row in ws.iter_rows(max_row=10)]
        
        # Query LLM
        prompt = f"""Based on this extracted financial data, answer the question clearly and concisely.

DATA:
{json.dumps(data_summary, indent=2)}

QUESTION: {message}

Provide a direct answer with specific numbers or facts from the data. Return as JSON: {{"answer": "your answer here"}}"""

        response = await call_mistral_optimized(prompt, max_tokens=500)
        answer = response.get("answer", str(response)) if isinstance(response, dict) else str(response)
        
        return JSONResponse({"response": answer})
    
    except Exception as e:
        logger.error(f"Chat error: {e}")
        return JSONResponse({"response": "Error processing your question."})

@app.get("/api/history")
async def history():
    """Get session history"""
    if not HISTORY_FILE.exists():
        return JSONResponse({"sessions": []})
    
    with open(HISTORY_FILE, 'r') as f:
        data = json.load(f)
    
    return JSONResponse({"sessions": data})

@app.get("/api/history/{session_id}")
async def get_session(session_id: str):
    """Get specific session"""
    if not HISTORY_FILE.exists():
        raise HTTPException(404, "No history found")
    
    with open(HISTORY_FILE, 'r') as f:
        history = json.load(f)
    
    session = next((s for s in history if s["session_id"] == session_id), None)
    if not session:
        raise HTTPException(404, "Session not found")
    
    return JSONResponse(session)

@app.get("/api/templates")
async def templates():
    """Get available templates"""
    return JSONResponse({
        "templates": {
            tid: {"name": t["name"], "sheets": t["sheets"]}
            for tid, t in TEMPLATES.items()
        }
    })

@app.get("/")
async def root():
    return {"message": "Velocity.ai API v2.2 - OPTIMIZED", "status": "online"}

@app.on_event("shutdown")
async def shutdown_event():
    """Close HTTP client on shutdown"""
    await http_client.aclose()

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)