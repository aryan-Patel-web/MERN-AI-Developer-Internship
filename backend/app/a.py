from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pathlib import Path
import pdfplumber, PyPDF2, json, uuid, logging, asyncio, os
from datetime import datetime
from typing import List, Dict
from mistralai.client import MistralClient
from groq import Groq
from dotenv import load_dotenv
from pdf2image import convert_from_path
import pytesseract
import openpyxl
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

# =========================
# Load Environment & Setup
# =========================
load_dotenv()
UPLOAD_DIR = Path("uploads"); UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR = Path("outputs"); OUTPUT_DIR.mkdir(exist_ok=True)
TEMPLATE_DIR = Path("templates"); TEMPLATE_DIR.mkdir(exist_ok=True)
HISTORY_FILE = Path("history.json"); HISTORY_FILE.touch(exist_ok=True)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("VelocityAI")

# =========================
# FastAPI App
# =========================
app = FastAPI(title="Velocity.ai - PDF Extraction Tool", version="1.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"], allow_credentials=True)

# =========================
# PDF Utilities
# =========================
def extract_pdf_text(file_path: Path) -> str:
    text = ""
    try:
        with pdfplumber.open(file_path) as pdf:
            for i, page in enumerate(pdf.pages, 1):
                page_text = page.extract_text() or ""
                if page_text.strip():
                    text += f"\n=== Page {i} ===\n{page_text}"
        if len(text.strip()) < 100:
            reader = PyPDF2.PdfReader(str(file_path))
            text = "\n".join([p.extract_text() or "" for p in reader.pages])
        if len(text.strip()) < 100:
            pages = convert_from_path(file_path)
            text = ""
            for i, page in enumerate(pages, 1):
                page_text = pytesseract.image_to_string(page)
                if page_text.strip():
                    text += f"\n=== Page {i} ===\n{page_text}"
    except Exception as e:
        logger.error(f"PDF extraction failed: {e}")
    return text

def is_scanned_pdf(file_path: Path) -> bool:
    try:
        reader = PyPDF2.PdfReader(str(file_path))
        total_text = "".join([p.extract_text() or "" for p in reader.pages])
        return len(total_text.strip()) < 50
    except:
        return True

# =========================
# Session Handling
# =========================
class Session(BaseModel):
    session_id: str
    created_at: str
    messages: List[dict] = []

def load_history():
    try:
        data = json.loads(HISTORY_FILE.read_text() or "[]")
        # Ensure it's a list of session dicts
        if not isinstance(data, list):
            data = []
        for i, s in enumerate(data):
            if not isinstance(s, dict) or "session_id" not in s:
                # Replace invalid entries with empty session dict
                data[i] = {
                    "session_id": str(uuid.uuid4()),
                    "created_at": datetime.now().isoformat(),
                    "messages": []
                }
        return data
    except Exception as e:
        logger.warning(f"Failed to load history: {e}")
        return []



def save_history(data):
    HISTORY_FILE.write_text(json.dumps(data, indent=2))

def add_session_message(session_id: str, message: dict):
    sessions = load_history()
    if not isinstance(sessions, list):
        sessions = []

    # Find existing session
    session_found = False
    for s in sessions:
        if isinstance(s, dict) and s.get("session_id") == session_id:
            s.setdefault("messages", []).append(message)
            session_found = True
            break

    # If session doesn't exist, create it
    if not session_found:
        new_session = {
            "session_id": session_id,
            "created_at": datetime.now().isoformat(),
            "messages": [message]
        }
        sessions.append(new_session)

    # Save back
    save_history(sessions)



# =========================
# LLM Service
# =========================
class LLMService:
    def __init__(self):
        self.mistral_api_key = os.getenv("MISTRAL_API_KEY")
        self.groq_api_key = os.getenv("GROQ_API_KEY")
        self.max_retries = 3
        self.mistral = MistralClient(api_key=self.mistral_api_key) if self.mistral_api_key else None
        self.groq = Groq(api_key=self.groq_api_key) if self.groq_api_key else None
        self.templates = self._load_templates()

    def _load_templates(self):
        templates = {}
        for t_file in TEMPLATE_DIR.glob("*.json"):
            with open(t_file) as f:
                templates[t_file.stem] = json.load(f)
        return templates

    def _build_prompt(self, text: str, template: dict) -> str:
        # STRICT prompt for real fields only

        
        return f"""
You are a highly accurate financial AI specializing in Private Equity fund data.

TASK:
- Extract ONLY the fields listed in the TEMPLATE (29 fields).
- Do NOT invent extra fields.
- Include 'value', 'confidence' (must be a number between 0 and 100), and 'source_page' for each field.
- If a value is missing (numeric or text), set value = "Not found" and confidence = 0.
- Preserve units for monetary values (75M, 980.0M) and percentages (25.0%).
- Include full page info in 'source_page' (e.g., "Page 1, opening letter").
- Include extra notes in 'source_page' where relevant (e.g., "Derived: No companies in this sector").
- Standardize all dates to "Month YYYY" (e.g., "December 2021").
- Remove duplicates from output (funds, companies).

DOCUMENT:
{text[:25000]}

TEMPLATE:
{json.dumps(template, indent=2)}

IMPORTANT:
- Return a SINGLE JSON object (dict), never a list.
- Top-level keys: fund_info, performance, portfolio, companies, metadata.
- Include all 29 template fields even if missing.
- Ensure no duplicate entries.

OUTPUT:
Return ONLY valid JSON matching the template structure with keys:
field, value, source_page, confidence.

"""

    async def extract_with_mistral(self, text: str, template: dict, retry=0):
        if not self.mistral:
            raise HTTPException(500, "Mistral not configured")
        try:
            prompt = self._build_prompt(text, template)
            response = self.mistral.chat(
                model="mistral-large-latest",
                messages=[{"role":"user","content":prompt}],
                temperature=0.1,
                max_tokens=4000
            )
            content = response.choices[0].message.content.strip()
            if content.startswith("```json"): content = content[7:]
            if content.endswith("```"): content = content[:-3]
            return json.loads(content)
        except json.JSONDecodeError:
            if retry < self.max_retries:
                await asyncio.sleep(2**retry)
                return await self.extract_with_mistral(text, template, retry+1)
            raise
        except Exception as e:
            logger.error(f"Mistral error: {e}")
            raise

    async def extract_with_groq(self, text: str, template: dict, retry=0):
        if not self.groq:
            raise HTTPException(500, "Groq not configured")
        try:
            prompt = self._build_prompt(text, template)
            response = self.groq.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[{"role": "user","content": prompt}],
                temperature=0.1, max_tokens=4000
            )
            content = response.choices[0].message.content.strip()
            if content.startswith("```json"): content = content[7:]
            if content.endswith("```"): content = content[:-3]
            return json.loads(content)
        except json.JSONDecodeError:
            if retry < self.max_retries:
                await asyncio.sleep(2**retry)
                return await self.extract_with_groq(text, template, retry+1)
            raise
        except Exception as e:
            logger.error(f"Groq error: {e}")
            raise

    async def extract(self, text: str, template_id: str):
        template = self.templates.get(template_id)
        if not template:
            raise HTTPException(400, f"Template '{template_id}' not found")
        try:
            return await self.extract_with_mistral(text, template)
        except Exception as e:
            logger.warning(f"Mistral failed, fallback to Groq: {e}")
        try:
            return await self.extract_with_groq(text, template)
        except Exception as e:
            raise HTTPException(500, "All LLM providers failed")

llm_service = LLMService()

# =========================
# Flattened JSON Helper
# =========================
def flatten_extracted_data(obj):
    """Flatten nested dict/list into a list of {field, value, source_page, confidence}"""
    flat = []
    if isinstance(obj, dict):
        if "field" in obj and "value" in obj:
            # normalize confidence
            conf = obj.get("confidence",0)
            if isinstance(conf,str): conf = conf.replace("%","").strip()
            try: conf = float(conf or 0)
            except: conf = 0
            flat.append({
                "field": obj.get("field",""),
                "value": obj.get("value",""),
                "source_page": obj.get("source_page",""),
                "confidence": conf
            })
        else:
            for v in obj.values():
                flat.extend(flatten_extracted_data(v))
    elif isinstance(obj, list):
        for item in obj:
            flat.extend(flatten_extracted_data(item))
    return flat

# =========================
# XLSX Export
# =========================
def save_to_xlsx(extracted_data: dict, output_file: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    flat_list = extracted_data.get("data_for_frontend", [])
    ws.append(["Field", "Value", "Source Page", "Confidence"])
    for row in flat_list:
        ws.append([row.get("field",""), row.get("value",""), row.get("source_page",""), row.get("confidence",0)])

    # adjust column width
    for col in ws.columns:
        max_length = max(len(str(cell.value) if cell.value else "") for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length+5,50)

    wb.save(output_file)

# =========================
# Process File
# =========================
async def process_file(file: UploadFile, template_id: str) -> Dict:
    job_file_id = str(uuid.uuid4())
    file_path = UPLOAD_DIR / f"{job_file_id}_{file.filename}"
    content = await file.read()
    file_path.write_bytes(content)

    text = extract_pdf_text(file_path)
    extraction_method = "OCR" if is_scanned_pdf(file_path) else "pdfplumber"

    data = await llm_service.extract(text, template_id)

    # Flatten for frontend & metadata
    flat_data = flatten_extracted_data(data)
    total_fields = len(flat_data)
    total_confidence = sum(f.get("confidence",0) for f in flat_data)
    avg_confidence = round(total_confidence/total_fields,3) if total_fields else 0

    data.setdefault("metadata",{})
    data["metadata"]["total_fields_extracted"] = total_fields
    data["metadata"]["average_confidence"] = avg_confidence
    data["data_for_frontend"] = flat_data

    output_file = OUTPUT_DIR / f"{file.filename.rsplit('.',1)[0]}_Extracted.xlsx"
    save_to_xlsx(data, output_file)

    return {
        "filename": file.filename,
        "status": "success",
        "data": data,
        "extraction_info": {"method": extraction_method, "char_count": len(text)},
        "llm_model": "mistral-large-latest" if llm_service.mistral else "llama-3.3-70b-versatile",
        "xlsx_path": str(output_file)
    }

# =========================
# API Endpoints
# =========================
@app.post("/api/extract")
async def extract(files: List[UploadFile] = File(...), template_id: str = Form("extraction_template_1"), session_id: str = Form(...)):
    tasks = [process_file(f, template_id) for f in files]
    results = await asyncio.gather(*tasks, return_exceptions=True)

    final_results = []
    for r in results:
        if isinstance(r, Exception):
            final_results.append({"status": "error", "error": str(r)})
        else:
            final_results.append(r)

    successful_results = [r for r in final_results if r.get("status")=="success"]
    total_fields = sum(r.get("data",{}).get("metadata",{}).get("total_fields_extracted",0) for r in successful_results)
    avg_confidence = round(sum(r.get("data",{}).get("metadata",{}).get("average_confidence",0) for r in successful_results)/max(1,len(successful_results)),3)

    summary_msg = {
        "role":"assistant",
        "content": f"Extraction completed: {len(successful_results)} successful, {len(final_results)-len(successful_results)} failed.",
        "timestamp": datetime.now().isoformat(),
        "results": final_results
    }
    add_session_message(session_id, summary_msg)

    summary = {
        "files_processed": len(files),
        "successful": len(successful_results),
        "failed": len(final_results)-len(successful_results),
        "total_fields_extracted": total_fields,
        "average_confidence": avg_confidence
    }

    return {"summary": summary, "results": final_results}

@app.get("/api/history")
def get_history():
    return {"sessions": load_history()}

@app.get("/api/history/{session_id}")
def get_session(session_id: str):
    sessions = load_history()
    for s in sessions:
        if s["session_id"]==session_id:
            return {"messages": s.get("messages",[])}
    return {"messages":[]}

@app.get("/api/download/{filename}")
async def download_file(filename: str):
    safe_file = OUTPUT_DIR / filename
    if not safe_file.exists():
        raise HTTPException(status_code=404, detail=f"File not found: {filename}")
    return FileResponse(
        path=safe_file,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        filename=filename
    )
