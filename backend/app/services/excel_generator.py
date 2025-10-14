import logging
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
from pathlib import Path

logger = logging.getLogger(__name__)

class ExcelGenerator:
    """
    Service for generating Excel files from extracted data following ILPA templates.
    """
    
    def __init__(self):
        self.templates = self._load_templates()
    
    def _load_templates(self) -> Dict[str, Any]:
        """Load Excel templates configuration."""
        templates = {}
        template_dir = Path("templates")
        
        for template_file in template_dir.glob("*.json"):
            try:
                with open(template_file, 'r') as f:
                    template_id = template_file.stem
                    templates[template_id] = json.load(f)
            except Exception as e:
                logger.error(f"Error loading template {template_file}: {e}")
        
        return templates
    
    def generate_excel(
        self,
        extracted_data: List[Dict[str, Any]],
        output_path: str,
        template_id: str
    ) -> str:
        """Generate Excel file from extracted data."""
        
        logger.info(f"Generating Excel file: {output_path}")
        
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        template = self.templates.get(template_id, {})
        sheets_config = template.get('excel_sheets', [])
        
        # Generate sheets based on template
        for sheet_config in sheets_config:
            sheet_name = sheet_config.get('name', 'Data')
            sheet_type = sheet_config.get('type', 'summary')
            
            if sheet_type == 'executive_summary':
                self._create_executive_summary(wb, extracted_data, sheet_config)
            elif sheet_type == 'schedule_of_investments':
                self._create_schedule_of_investments(wb, extracted_data, sheet_config)
            elif sheet_type == 'portfolio_companies':
                self._create_portfolio_companies(wb, extracted_data, sheet_config)
            elif sheet_type == 'financial_statements':
                self._create_financial_statements(wb, extracted_data, sheet_config)
            elif sheet_type == 'footnotes':
                self._create_footnotes(wb, extracted_data, sheet_config)
            else:
                self._create_generic_sheet(wb, extracted_data, sheet_config)
        
        # Save workbook
        wb.save(output_path)
        logger.info(f"Excel file generated successfully: {output_path}")
        
        return output_path
    
    def _create_executive_summary(
        self,
        wb: Workbook,
        data: List[Dict[str, Any]],
        config: Dict[str, Any]
    ):
        """Create Executive Summary sheet."""
        
        ws = wb.create_sheet(config.get('name', 'Executive Summary'))
        
        # Merge data from all files
        merged_data = self._merge_fund_data(data)
        
        # Header
        ws['A1'] = 'Fund Executive Summary'
        ws['A1'].font = Font(size=16, bold=True)
        
        row = 3
        
        # General Partner Info
        ws[f'A{row}'] = 'General Partner:'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = merged_data.get('general_partner', 'N/A')
        row += 2
        
        # Fund Details
        fund_fields = [
            ('Fund Name', 'fund_name'),
            ('Fund Currency', 'fund_currency'),
            ('Assets Under Management', 'assets_under_management'),
            ('Active Funds', 'active_funds'),
            ('Active Portfolio Companies', 'active_portfolio_companies'),
            ('Total Commitments', 'total_commitments'),
            ('Total Drawdowns', 'total_drawdowns'),
            ('Remaining Commitments', 'remaining_commitments'),
            ('Total Number of Investments', 'total_number_of_investments'),
            ('Total Distributions', 'total_distributions'),
        ]
        
        for label, field in fund_fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = merged_data.get(field, 'N/A')
            row += 1
        
        row += 1
        
        # Financial Metrics
        ws[f'A{row}'] = 'Key Financial Metrics'
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        metrics = [
            ('DPI (Distribution to Paid-in Capital)', 'dpi'),
            ('RVPI (Residual Value to Paid-in Capital)', 'rvpi'),
            ('TVPI (Total Value to Paid-in Capital)', 'tvpi'),
            ('Net IRR', 'net_irr'),
            ('Gross IRR', 'gross_irr'),
        ]
        
        for label, field in metrics:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = merged_data.get(field, 'N/A')
            row += 1
        
        # Auto-adjust column width
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 25
    
    def _create_schedule_of_investments(
        self,
        wb: Workbook,
        data: List[Dict[str, Any]],
        config: Dict[str, Any]
    ):
        """Create Schedule of Investments sheet."""
        
        ws = wb.create_sheet(config.get('name', 'Schedule of Investments'))
        
        # Headers
        headers = [
            'Company Name',
            'Security Type',
            'Number of Shares',
            'Fund Ownership %',
            'Initial Investment Date',
            'Fund Commitment',
            'Total Invested',
            'Current Cost',
            'Reported Value',
            'Realized Proceeds',
            'Multiple of Invested Capital',
            'Gross IRR'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Data rows
        row = 2
        for file_data in data:
            investments = file_data.get('data', {}).get('investments', [])
            
            for investment in investments:
                ws.cell(row=row, column=1, value=investment.get('company_name', ''))
                ws.cell(row=row, column=2, value=investment.get('security_type', ''))
                ws.cell(row=row, column=3, value=investment.get('number_of_shares', ''))
                ws.cell(row=row, column=4, value=investment.get('fund_ownership_percent', ''))
                ws.cell(row=row, column=5, value=investment.get('initial_investment_date', ''))
                ws.cell(row=row, column=6, value=investment.get('fund_commitment', ''))
                ws.cell(row=row, column=7, value=investment.get('total_invested', ''))
                ws.cell(row=row, column=8, value=investment.get('current_cost', ''))
                ws.cell(row=row, column=9, value=investment.get('reported_value', ''))
                ws.cell(row=row, column=10, value=investment.get('realized_proceeds', ''))
                ws.cell(row=row, column=11, value=investment.get('multiple_of_invested_capital', ''))
                ws.cell(row=row, column=12, value=investment.get('gross_irr', ''))
                row += 1
        
        # Format columns
        for col in range(1, 13):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def _create_portfolio_companies(
        self,
        wb: Workbook,
        data: List[Dict[str, Any]],
        config: Dict[str, Any]
    ):
        """Create Portfolio Companies sheet with detailed company information."""
        
        ws = wb.create_sheet(config.get('name', 'Portfolio Companies'))
        
        row = 1
        
        for file_data in data:
            companies = file_data.get('data', {}).get('portfolio_companies', [])
            
            for company in companies:
                # Company header
                ws[f'A{row}'] = company.get('company_name', 'Unknown Company')
                ws[f'A{row}'].font = Font(size=14, bold=True, color="FFFFFF")
                ws[f'A{row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                ws.merge_cells(f'A{row}:D{row}')
                row += 1
                
                # Company details
                company_fields = [
                    ('Address', 'address'),
                    ('Website', 'website'),
                    ('Business Sector', 'business_sector'),
                    ('Primary Industry', 'primary_industry'),
                    ('Initial Investment Stage', 'initial_investment_stage'),
                    ('Region', 'region'),
                    ('Date of Investment', 'date_of_investment'),
                    ('Invested Capital', 'invested_capital'),
                    ('Currency', 'currency'),
                    ('Capital Returned', 'capital_returned'),
                    ('Location', 'location'),
                    ('Ownership %', 'ownership_percent'),
                    ('Investment Type', 'investment_type'),
                    ('Latest Audited FY', 'latest_audited_fy'),
                    ('Latest Quarter', 'latest_quarter'),
                ]
                
                for label, field in company_fields:
                    ws[f'A{row}'] = label
                    ws[f'A{row}'].font = Font(bold=True)
                    ws[f'B{row}'] = company.get(field, 'N/A')
                    row += 1
                
                # Company description
                if 'company_description' in company:
                    ws[f'A{row}'] = 'Company Description'
                    ws[f'A{row}'].font = Font(bold=True)
                    row += 1
                    ws[f'A{row}'] = company.get('company_description', '')
                    ws[f'A{row}'].alignment = Alignment(wrap_text=True)
                    ws.merge_cells(f'A{row}:D{row}')
                    row += 1
                
                # Investment thesis
                if 'investment_thesis' in company:
                    ws[f'A{row}'] = 'Investment Thesis'
                    ws[f'A{row}'].font = Font(bold=True)
                    row += 1
                    ws[f'A{row}'] = company.get('investment_thesis', '')
                    ws[f'A{row}'].alignment = Alignment(wrap_text=True)
                    ws.merge_cells(f'A{row}:D{row}')
                    row += 1
                
                # Historical performance
                if 'historical_performance' in company:
                    ws[f'A{row}'] = 'Historical Performance'
                    ws[f'A{row}'].font = Font(bold=True, size=12)
                    row += 1
                    
                    perf_data = company.get('historical_performance', {})
                    years = perf_data.get('years', [])
                    
                    if years:
                        # Headers
                        ws[f'A{row}'] = 'Metric'
                        for i, year in enumerate(years, 2):
                            ws.cell(row=row, column=i, value=year)
                        row += 1
                        
                        # Revenue
                        ws[f'A{row}'] = 'Revenue'
                        revenues = perf_data.get('revenue', [])
                        for i, rev in enumerate(revenues, 2):
                            ws.cell(row=row, column=i, value=rev)
                        row += 1
                        
                        # EBITDA
                        ws[f'A{row}'] = 'EBITDA'
                        ebitdas = perf_data.get('ebitda', [])
                        for i, ebitda in enumerate(ebitdas, 2):
                            ws.cell(row=row, column=i, value=ebitda)
                        row += 1
                        
                        # EBITDA Margin
                        ws[f'A{row}'] = 'EBITDA Margin %'
                        margins = perf_data.get('ebitda_margin', [])
                        for i, margin in enumerate(margins, 2):
                            ws.cell(row=row, column=i, value=margin)
                        row += 1
                
                # Recent performance
                if 'recent_performance' in company:
                    ws[f'A{row}'] = 'Recent Performance'
                    ws[f'A{row}'].font = Font(bold=True, size=12)
                    row += 1
                    
                    recent = company.get('recent_performance', {})
                    
                    recent_fields = [
                        ('Quarter', 'quarter'),
                        ('Revenue', 'revenue'),
                        ('Revenue Change vs PY', 'revenue_change'),
                        ('EBITDA', 'ebitda'),
                        ('EBITDA Change vs PY', 'ebitda_change'),
                        ('EBITDA Margin %', 'ebitda_margin'),
                    ]
                    
                    for label, field in recent_fields:
                        ws[f'A{row}'] = label
                        ws[f'B{row}'] = recent.get(field, 'N/A')
                        row += 1
                
                row += 2  # Space between companies
        
        # Format columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
    
    def _create_financial_statements(
        self,
        wb: Workbook,
        data: List[Dict[str, Any]],
        config: Dict[str, Any]
    ):
        """Create Financial Statements sheets."""
        
        # Income Statement
        ws_income = wb.create_sheet('Statement of Operations')
        self._populate_income_statement(ws_income, data)
        
        # Balance Sheet
        ws_balance = wb.create_sheet('Balance Sheet')
        self._populate_balance_sheet(ws_balance, data)
        
        # Cash Flow
        ws_cashflow = wb.create_sheet('Statement of Cash Flows')
        self._populate_cashflow_statement(ws_cashflow, data)
        
        # Partners Capital
        ws_pcap = wb.create_sheet('Partners Capital Statement')
        self._populate_pcap_statement(ws_pcap, data)
    
    def _populate_income_statement(self, ws, data):
        """Populate income statement data."""
        
        merged_data = self._merge_fund_data(data)
        income_stmt = merged_data.get('income_statement', {})
        
        ws['A1'] = 'Statement of Operations'
        ws['A1'].font = Font(size=14, bold=True)
        
        row = 3
        
        # Income section
        ws[f'A{row}'] = 'Income'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        income_items = [
            ('Dividend Income', 'dividend_income'),
            ('Interest Income', 'interest_income'),
            ('Other Income', 'other_income'),
        ]
        
        total_income = 0
        for label, field in income_items:
            value = income_stmt.get(field, 0)
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            if isinstance(value, (int, float)):
                total_income += value
            row += 1
        
        ws[f'A{row}'] = 'Total Income'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = total_income
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        # Expenses section
        ws[f'A{row}'] = 'Expenses'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        expense_items = [
            ('Management Fees', 'management_fees'),
            ('Fund Expenses', 'fund_expenses'),
            ('Other Expenses', 'other_expenses'),
        ]
        
        total_expenses = 0
        for label, field in expense_items:
            value = income_stmt.get(field, 0)
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            if isinstance(value, (int, float)):
                total_expenses += value
            row += 1
        
        ws[f'A{row}'] = 'Total Expenses'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = total_expenses
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        # Net unrealized gain
        ws[f'A{row}'] = 'Net Unrealized Gain on Investments'
        ws[f'B{row}'] = income_stmt.get('net_unrealized_gain', 0)
        row += 1
        
        # Total comprehensive income
        comprehensive_income = total_income - total_expenses + income_stmt.get('net_unrealized_gain', 0)
        ws[f'A{row}'] = 'Total Comprehensive Income'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'] = comprehensive_income
        ws[f'B{row}'].font = Font(bold=True)
        
        # Format
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
    
    def _populate_balance_sheet(self, ws, data):
        """Populate balance sheet data."""
        
        merged_data = self._merge_fund_data(data)
        balance_sheet = merged_data.get('balance_sheet', {})
        
        ws['A1'] = 'Balance Sheet'
        ws['A1'].font = Font(size=14, bold=True)
        
        row = 3
        
        # Assets
        ws[f'A{row}'] = 'ASSETS'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        asset_items = [
            ('Investments at Fair Value', 'investments_fair_value'),
            ('Cash and Cash Equivalents', 'cash_and_equivalents'),
            ('Amount Due from Limited Partners', 'due_from_lps'),
            ('Other Assets', 'other_assets'),
        ]
        
        total_assets = 0
        for label, field in asset_items:
            value = balance_sheet.get(field, 0)
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            if isinstance(value, (int, float)):
                total_assets += value
            row += 1
        
        ws[f'A{row}'] = 'Total Assets'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = total_assets
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        # Liabilities
        ws[f'A{row}'] = 'LIABILITIES'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        liability_items = [
            ('Amount Due to Related Party', 'due_to_related_party'),
            ('Other Payables', 'other_payables'),
        ]
        
        total_liabilities = 0
        for label, field in liability_items:
            value = balance_sheet.get(field, 0)
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            if isinstance(value, (int, float)):
                total_liabilities += value
            row += 1
        
        ws[f'A{row}'] = 'Total Liabilities'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = total_liabilities
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        # Partners Capital
        ws[f'A{row}'] = "Partners' Capital"
        ws[f'B{row}'] = total_assets - total_liabilities
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'].font = Font(bold=True)
        
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
    
    def _populate_cashflow_statement(self, ws, data):
        """Populate cash flow statement."""
        
        ws['A1'] = 'Statement of Cash Flows'
        ws['A1'].font = Font(size=14, bold=True)
        ws['A3'] = 'Cash flow data extracted from financial statements'
        
        ws.column_dimensions['A'].width = 40
    
    def _populate_pcap_statement(self, ws, data):
        """Populate partners capital account statement."""
        
        merged_data = self._merge_fund_data(data)
        pcap = merged_data.get('partners_capital', {})
        
        ws['A1'] = "Partners' Capital Account Statement"
        ws['A1'].font = Font(size=14, bold=True)
        
        row = 3
        
        pcap_items = [
            ('Total Commitments', 'total_commitments'),
            ('Contributed Capital', 'contributed_capital'),
            ('Net Operating Loss', 'net_operating_loss'),
            ('Total Partners Capital', 'total_partners_capital'),
        ]
        
        for label, field in pcap_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = pcap.get(field, 'N/A')
            row += 1
        
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
    
    def _create_footnotes(
        self,
        wb: Workbook,
        data: List[Dict[str, Any]],
        config: Dict[str, Any]
    ):
        """Create footnotes sheet."""
        
        ws = wb.create_sheet(config.get('name', 'Footnotes'))
        
        ws['A1'] = 'Footnotes and Disclosures'
        ws['A1'].font = Font(size=14, bold=True)
        
        row = 3
        
        merged_data = self._merge_fund_data(data)
        footnotes = merged_data.get('footnotes', [])
        
        for idx, footnote in enumerate(footnotes, 1):
            ws[f'A{row}'] = f"Note {idx}"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            ws[f'A{row}'] = footnote.get('title', '')
            ws[f'A{row}'].font = Font(italic=True)
            row += 1
            
            ws[f'A{row}'] = footnote.get('content', '')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            row += 2
        
        ws.column_dimensions['A'].width = 80
    
    def _create_generic_sheet(
        self,
        wb: Workbook,
        data: List[Dict[str, Any]],
        config: Dict[str, Any]
    ):
        """Create a generic data sheet."""
        
        ws = wb.create_sheet(config.get('name', 'Data'))
        
        ws['A1'] = config.get('title', 'Extracted Data')
        ws['A1'].font = Font(size=14, bold=True)
        
        # Simple key-value layout
        row = 3
        for file_data in data:
            ws[f'A{row}'] = 'Source File'
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = file_data.get('filename', '')
            row += 2
            
            extracted = file_data.get('data', {})
            for key, value in extracted.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = str(value)
                row += 1
            
            row += 2
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50
    
    def _merge_fund_data(self, data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Merge data from multiple PDF files into a single fund view."""
        
        merged = {}
        
        for file_data in data:
            extracted = file_data.get('data', {})
            for key, value in extracted.items():
                if key not in merged:
                    merged[key] = value
                elif isinstance(value, list):
                    if key not in merged:
                        merged[key] = []
                    merged[key].extend(value)
        
        return merged